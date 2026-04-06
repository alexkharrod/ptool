"""
Build a preview spreadsheet of the first 5 PromoStandards products.
Run: python build_ps_preview.py
"""
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── Data fetched live from LogoIncluded PromoStandards API ────────────────────
PRODUCTS = [
    {
        "productId":      "EC18",
        "productName":    "Slim Credit Card Style Apple AirTag Holder",
        "category":       "Personal Tech",
        "description":    (
            "The Credit Card Style AirTag holder is slim and lightweight, so it easily slides "
            "into your wallet. It protects your AirTag from scratches and fits securely so "
            "there's no need to worry about it falling out. Ideal for a purse, backpack, "
            "luggage, etc. The open-hole design in the top corner can be used with key rings "
            "or a carabiner. Composed of silicone — durable and easy to clean."
        ),
        "colors":         "Black, White",
        "partIds":        "EC18_Black, EC18_White",
        "moq":            250,
        "priceTiers":     "250: $4.62 | 500: $4.16 | 1000: $3.78",
        "imprint":        "front",
        "dimensions":     "2.0\" × 0.1\" IN",
        "imageUrl":       "https://d207zvy2rsg5b5.cloudfront.net/B0CDADFD-F477-4E18-94A3-110C470D6097/Products/Large/EC18-6-Th.jpg",
        "countryOfOrigin":"US",
    },
    {
        "productId":      "TT01",
        "productName":    "Twiddle-Twitch Fidget Cube",
        "category":       "Personal Tech",
        "description":    (
            "The Twiddle-Twitch is perfect for people who can't keep still! Designed to help "
            "reduce stress and anxiety, this device features 6 different sides: an analog stick, "
            "a switch, a spinner, a clicker, 5 buttons and a flat imprint side. Made of durable "
            "plastic with a soft finish. Lightweight and compact. *Not a children's toy — not "
            "intended for children under age 12."
        ),
        "colors":         "Black, White, Blue, Light Green, Orange, Red Dark",
        "partIds":        "TT01_Black, TT01_White, TT01_Blue, TT01_Light Green, TT01_Orange, TT01_Red Dark",
        "moq":            100,
        "priceTiers":     "100: $4.29 | 250: $3.86 | 500: $3.55 | 1000: $3.33",
        "imprint":        "side; front side",
        "dimensions":     "1.25\" × 1.25\" × 1.25\" IN",
        "imageUrl":       "https://d207zvy2rsg5b5.cloudfront.net/B0CDADFD-F477-4E18-94A3-110C470D6097/Products/Large/TT01wlogo-Th.jpg",
        "countryOfOrigin":"US",
    },
    {
        "productId":      "BA03-US",
        "productName":    "Terra Eco Friendly USB Data Blocker",
        "category":       "Personal Tech",
        "description":    (
            "LAST CHANCE PRICING — The Terra Eco Friendly Data Blocker protects your devices "
            "from viruses and malware while charging. Made with a wheat straw shell for a more "
            "sustainable alternative. Compact design — easy to carry or attach to keys/bags. "
            "Compatible with Apple, Micro-USB, and Type-C devices. Ensures safe, charge-only "
            "connections when plugged into any power source."
        ),
        "colors":         "Default",
        "partIds":        "BA03-US",
        "moq":            100,
        "priceTiers":     "100: $3.35 | 250: $3.35",
        "imprint":        "front",
        "dimensions":     "1.7\" × 0.3\" IN",
        "imageUrl":       "https://d207zvy2rsg5b5.cloudfront.net/B0CDADFD-F477-4E18-94A3-110C470D6097/Products/Large/BA03-US-Th.jpg",
        "countryOfOrigin":"US",
    },
    {
        "productId":      "IC33",
        "productName":    "10.9 in. iPad Air 5th Generation Case",
        "category":       "Mobile Tech",
        "description":    (
            "Preserve your device with a protective case! The 10.9\" case provides a snug fit "
            "for your 4th or 5th generation iPad Air with a hard interior cover and premium "
            "synthetic leather exterior. Protects from scratches and harmful elements. Rotating "
            "structure allows flexible horizontal and vertical viewing. 360-degree rotation with "
            "three prop-up viewing angles. Great for business meetings, travel, school and more!"
        ),
        "colors":         "Black",
        "partIds":        "IC33_Black",
        "moq":            50,
        "priceTiers":     "50: $12.37 | 100: $11.88 | 250: $11.65 | 500: $11.43 | 1000: $11.02",
        "imprint":        "back of case; front case",
        "dimensions":     "7.2\" × 0.5\" IN",
        "imageUrl":       "https://d207zvy2rsg5b5.cloudfront.net/B0CDADFD-F477-4E18-94A3-110C470D6097/Products/Large/518c1f94-1865-41da-9cc0-d527f587ed73.jpg",
        "countryOfOrigin":"US",
    },
    {
        "productId":      "LY1801",
        "productName":    "1 in Wrist Lanyard with Full Color Sublimation Imprint",
        "category":       "Lanyards / Sublimated",
        "description":    (
            "A Wrist Lanyard is great for ID badges, keys, and makes a great giveaway item! "
            "This 1\" wristband is fully sublimated on both sides — ideal for any step-and-repeat "
            "logo. Use for security/pass reasons at events or to identify luggage. A great way "
            "to stand out in a crowd and brand your logo. Very comfortable material."
        ),
        "colors":         "Any Color — Completely Custom",
        "partIds":        "LY1801_Any Color - Completely Custom",
        "moq":            500,
        "priceTiers":     "500: $0.70 | 1000: $0.62",
        "imprint":        "fully sublimated (both sides)",
        "dimensions":     "15.7\" × 1.0\" IN",
        "imageUrl":       "https://d207zvy2rsg5b5.cloudfront.net/B0CDADFD-F477-4E18-94A3-110C470D6097/Products/Large/89a9e86e-58e7-47d5-a60f-7182ea06466e.jpg",
        "countryOfOrigin":"US",
    },
]

# ── Build workbook ─────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "PS Product Preview"

# Styles
HEADER_BG   = "1F3864"   # dark navy
HEADER_FG   = "FFFFFF"
TITLE_BG    = "2E75B6"   # LogoIncluded blue
TITLE_FG    = "FFFFFF"
ALT_BG      = "EBF3FB"   # light blue alternating row
WHITE       = "FFFFFF"
BORDER_COLOR= "CCCCCC"

thin = Side(style="thin", color=BORDER_COLOR)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hfill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def cell_style(ws, row, col, value="", bold=False, bg=None, fg="000000",
               wrap=True, align="left", valign="top", border=None, hyperlink=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=fg, name="Arial", size=10)
    c.alignment = Alignment(
        horizontal=align, vertical=valign,
        wrap_text=wrap
    )
    if bg:
        c.fill = hfill(bg)
    if border:
        c.border = border
    if hyperlink:
        c.hyperlink = hyperlink
        c.font = Font(bold=bold, color="0563C1", name="Arial", size=10, underline="single")
    return c

# ── Title row ─────────────────────────────────────────────────────────────────
ws.merge_cells("A1:M1")
title = ws["A1"]
title.value = "LogoIncluded PromoStandards API — Product Preview (First 5 Products)"
title.font = Font(bold=True, color=TITLE_FG, name="Arial", size=13)
title.fill = hfill(TITLE_BG)
title.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

ws.merge_cells("A2:M2")
sub = ws["A2"]
sub.value = "Source: LogoIncluded PromoStandards API  |  getProduct + getConfigurationAndPricing  |  April 6, 2026"
sub.font = Font(italic=True, color="555555", name="Arial", size=9)
sub.fill = hfill("D6E4F0")
sub.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 16

# ── Column headers ─────────────────────────────────────────────────────────────
HEADERS = [
    "Product ID",
    "Product Name",
    "Category",
    "Description",
    "Colors Available",
    "Part IDs",
    "MOQ",
    "Price Tiers (USD)",
    "Imprint Location",
    "Dimensions",
    "Country of Origin",
    "Image URL",
    "Notes",
]

HEADER_ROW = 3
for col, h in enumerate(HEADERS, 1):
    cell_style(ws, HEADER_ROW, col, h,
               bold=True, bg=HEADER_BG, fg=HEADER_FG,
               align="center", valign="center", wrap=True)

ws.row_dimensions[HEADER_ROW].height = 32

# ── Data rows ─────────────────────────────────────────────────────────────────
for i, p in enumerate(PRODUCTS):
    row = HEADER_ROW + 1 + i
    bg = ALT_BG if i % 2 == 0 else WHITE

    cell_style(ws, row, 1,  p["productId"],      bold=True, bg=bg)
    cell_style(ws, row, 2,  p["productName"],     bold=False, bg=bg)
    cell_style(ws, row, 3,  p["category"],        bg=bg)
    cell_style(ws, row, 4,  p["description"],     bg=bg, wrap=True)
    cell_style(ws, row, 5,  p["colors"],          bg=bg)
    cell_style(ws, row, 6,  p["partIds"],         bg=bg)
    cell_style(ws, row, 7,  p["moq"],             bg=bg, align="center")
    cell_style(ws, row, 8,  p["priceTiers"],      bg=bg)
    cell_style(ws, row, 9,  p["imprint"],         bg=bg)
    cell_style(ws, row, 10, p["dimensions"],      bg=bg, align="center")
    cell_style(ws, row, 11, p["countryOfOrigin"], bg=bg, align="center")
    cell_style(ws, row, 12, "View Image", bg=bg, hyperlink=p["imageUrl"])
    cell_style(ws, row, 13, "",                   bg=bg)

    ws.row_dimensions[row].height = 80

# ── Column widths ──────────────────────────────────────────────────────────────
COL_WIDTHS = {
    1: 12,   # Product ID
    2: 32,   # Name
    3: 18,   # Category
    4: 55,   # Description
    5: 30,   # Colors
    6: 38,   # Part IDs
    7: 8,    # MOQ
    8: 40,   # Price Tiers
    9: 28,   # Imprint
    10: 18,  # Dimensions
    11: 10,  # Country
    12: 12,  # Image URL
    13: 20,  # Notes
}
for col, width in COL_WIDTHS.items():
    ws.column_dimensions[get_column_letter(col)].width = width

# ── Freeze panes below header ──────────────────────────────────────────────────
ws.freeze_panes = "A4"

# ── Auto-filter on header row ──────────────────────────────────────────────────
ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(len(HEADERS))}{HEADER_ROW + len(PRODUCTS)}"

OUT = "/sessions/sweet-adoring-hawking/mnt/ptool/ps_product_preview.xlsx"
wb.save(OUT)
print(f"Saved: {OUT}")
