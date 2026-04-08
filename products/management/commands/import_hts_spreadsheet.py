"""
Management command: import_hts_spreadsheet

Reads the filled-in hts_codes.xlsx and:
  1. Creates any missing Category records
  2. Creates any missing HtsCode records (duty rates = 0 — fill in via HTS Codes admin)
  3. Updates every product: category (CharField) + hts_code (FK)

Usage:
    python manage.py import_hts_spreadsheet
    python manage.py import_hts_spreadsheet --input hts_codes.xlsx
    python manage.py import_hts_spreadsheet --dry-run
"""

import os

from django.core.management.base import BaseCommand
from django.db import transaction

from products.models import Category, HtsCode, Product

# Best-guess descriptions for category codes derived from SKU prefixes.
# Edit directly in the admin after import if any are wrong.
CATEGORY_DESCRIPTIONS = {
    "AC":   "AC Adapters",
    "AT":   "Air Trackers",
    "BO":   "Bottle Openers",
    "CA":   "Car Adapters",
    "CB":   "Charging Cables",
    "CL":   "Camera Lenses",
    "CM":   "Custom Molded",
    "CR":   "Croc Charms",
    "CS":   "Cases",
    "CT":   "Cooling Towels",
    "DF":   "Digital Photo Frames",
    "DM":   "Distance Measurers",
    "DW":   "Drinkware",
    "EB":   "Earbuds",
    "EC":   "Electronics Cases",
    "ES":   "Electronics / A/V",
    "FN":   "Fans",
    "FT":   "Fitness Accessories",
    "HW":   "Hand Warmers",
    "IC":   "iPad / Tablet Cases",
    "IS":   "iPhone / Phone Stands",
    "JB":   "Power Banks",
    "KC":   "Keychains",
    "KT":   "Kits",
    "LG":   "Luggage / Badge Accessories",
    "LY":   "Lanyards",
    "MA":   "Microphones / Audio Accessories",
    "MB":   "Magic Boards",
    "MG":   "Magnetic Accessories",
    "Misc": "Miscellaneous",
    "MP":   "Mouse Pads",
    "MR":   "Mice / Input Devices",
    "NFC":  "NFC Tags",
    "OA":   "Office Accessories",
    "OM":   "Mice / Optical Input",
    "PH":   "Phone Holders",
    "PK":   "Packaging",
    "PM":   "Presenters / Pointers",
    "PN":   "Pens",
    "PW":   "Phone Wallets",
    "RT":   "Retail",
    "SA":   "Safety Alarms",
    "SC":   "Screen Cleaners",
    "SL":   "Sleeves / Lanyards",
    "SLV":  "Card Sleeves",
    "SP":   "Speakers",
    "ST":   "Straws",
    "TA":   "Travel Adapters",
    "TL":   "Tools / Lights",
    "TS":   "Styluses",
    "TT":   "Toys / Trinkets",
    "UA":   "USB Accessories",
    "UD":   "USB Drives",
    "UH":   "USB Hubs",
    "UR":   "USB Readers",
    "WB":   "Waterproof Bags",
    "WC":   "Wireless Chargers",
    "WR":   "Wristbands",
}

# HTS codes to create: description, duty%, section_301%, extra_tariff%
NEW_HTS_CODES = {
    "7323.93.0080": ("Table, kitchen or other household articles of stainless steel (other)", 2.0,  25.0, 10.0),
    "7326.90.8600": ("Other articles of iron or steel (other)",                                2.9,  25.0, 10.0),
    "8471.90.0000": ("Other automatic data-processing machines and units thereof",             0.0,  25.0, 10.0),
    "8509.40.1000": ("Electromechanical domestic appliances — food grinders, mixers",          3.4,  25.0, 10.0),
    "8518.10.8000": ("Microphones (other)",                                                    0.0,   0.0, 10.0),
}


class Command(BaseCommand):
    help = "Import categories and HTS codes from the filled-in hts_codes.xlsx and update all products"

    def add_arguments(self, parser):
        parser.add_argument(
            "--input", default="hts_codes.xlsx",
            help="Path to the filled-in spreadsheet (default: hts_codes.xlsx)",
        )
        parser.add_argument(
            "--dry-run", action="store_true",
            help="Preview changes without saving",
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            self.stderr.write(self.style.ERROR("openpyxl not installed. Run: pip install openpyxl"))
            return

        input_path = options["input"]
        dry_run    = options["dry_run"]

        if not os.path.exists(input_path):
            self.stderr.write(self.style.ERROR(f"File not found: {input_path}"))
            return

        if dry_run:
            self.stdout.write(self.style.WARNING("DRY RUN — nothing will be saved\n"))

        # ── 1. Read spreadsheet ───────────────────────────────────────────
        wb = openpyxl.load_workbook(input_path, data_only=True)
        ws = wb["Products"]

        rows = []
        for row in range(2, ws.max_row + 1):
            sku     = ws.cell(row, 1).value
            cat     = ws.cell(row, 3).value
            hts_raw = ws.cell(row, 5).value
            if not sku:
                continue
            rows.append({
                "sku":      str(sku).strip().upper(),
                "category": str(cat).strip() if cat else "",
                "hts_code": str(hts_raw).strip() if hts_raw and str(hts_raw).strip().upper() != "N/A" else "",
            })

        self.stdout.write(f"Read {len(rows)} product rows from spreadsheet")

        # ── 2. Collect unique values ──────────────────────────────────────
        needed_cats = {r["category"] for r in rows if r["category"]}
        needed_hts  = {r["hts_code"] for r in rows if r["hts_code"]}

        existing_cats = set(Category.objects.values_list("code", flat=True))
        existing_hts  = set(HtsCode.objects.values_list("code", flat=True))

        cats_to_create = needed_cats - existing_cats
        hts_to_create  = needed_hts  - existing_hts

        self.stdout.write(f"\nCategories: {len(needed_cats)} needed, "
                          f"{len(existing_cats & needed_cats)} exist, "
                          f"{len(cats_to_create)} to create")
        for code in sorted(cats_to_create):
            desc = CATEGORY_DESCRIPTIONS.get(code, f"{code} Products")
            self.stdout.write(f"  + Category  {code:10s}  {desc}")

        self.stdout.write(f"\nHTS Codes: {len(needed_hts)} needed, "
                          f"{len(existing_hts & needed_hts)} exist, "
                          f"{len(hts_to_create)} to create")
        for code in sorted(hts_to_create):
            desc, duty, s301, extra = NEW_HTS_CODES.get(code, ("(description not set)", 0, 0, 0))
            self.stdout.write(f"  + HTS Code  {code}  {desc}  ({duty}% / {s301}% / {extra}%)")

        if dry_run:
            self.stdout.write(self.style.WARNING(
                f"\nDRY RUN — would update {len(rows)} products"
            ))
            return

        # ── 3. Create missing categories ──────────────────────────────────
        with transaction.atomic():
            for code in sorted(cats_to_create):
                desc = CATEGORY_DESCRIPTIONS.get(code, f"{code} Products")
                Category.objects.create(code=code, description=desc)
            if cats_to_create:
                self.stdout.write(self.style.SUCCESS(
                    f"\nCreated {len(cats_to_create)} categories ✓"
                ))

            # ── 4. Create missing HTS codes ───────────────────────────────
            for code in sorted(hts_to_create):
                desc, duty, s301, extra = NEW_HTS_CODES.get(code, ("(description not set)", 0, 0, 0))
                HtsCode.objects.create(
                    code=code,
                    description=desc,
                    duty_percent=duty,
                    section_301_percent=s301,
                    extra_tariff_percent=extra,
                )
            if hts_to_create:
                self.stdout.write(self.style.SUCCESS(
                    f"Created {len(hts_to_create)} HTS codes ✓"
                ))

            # ── 5. Build lookup caches ────────────────────────────────────
            hts_map = {h.code: h for h in HtsCode.objects.all()}

            # ── 6. Update products ────────────────────────────────────────
            updated = skipped = errors = 0
            for r in rows:
                try:
                    product = Product.objects.get(sku=r["sku"])
                    product.category = r["category"]
                    hts_obj = hts_map.get(r["hts_code"]) if r["hts_code"] else None
                    product.hts_code = hts_obj
                    # Sync rate fields from HTS code
                    product.duty_percent   = hts_obj.duty_percent          if hts_obj else 0
                    product.tariff_percent = hts_obj.section_301_percent   if hts_obj else 0
                    product.save(update_fields=["category", "hts_code", "duty_percent", "tariff_percent"])
                    updated += 1
                except Product.DoesNotExist:
                    self.stderr.write(f"  SKU not found: {r['sku']}")
                    skipped += 1
                except Exception as e:
                    self.stderr.write(self.style.ERROR(f"  Error on {r['sku']}: {e}"))
                    errors += 1

        self.stdout.write("\n" + "─" * 40)
        self.stdout.write(self.style.SUCCESS("Import complete"))
        self.stdout.write(f"  Products updated : {updated}")
        self.stdout.write(f"  SKUs not found   : {skipped}")
        self.stdout.write(f"  Errors           : {errors}")
        self.stdout.write(
            "\nReminder: Set duty rates on new HTS codes at /hts-codes/ in the app."
        )
