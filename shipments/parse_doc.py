"""
Packing-list / commercial-invoice parser for uploaded XLS and XLSX files.

Supports two common supplier layouts:
  - Packing list sheet: rows with PO#, P/N, description, cartons, qty, NW, GW, dims, CBM
  - CI sheet: rows with qty, PO#, product#, description, unit cost

Returns a dict:
    {
        "items": [
            {
                "po_number": "81300",
                "sku": "20-3008-1085",
                "description": "AD Player",
                "cartons": 20,
                "qty": 200,
                "nw_kg": "280.00",
                "gw_kg": "318.00",
                "dimensions_cm": "42×36×39.5",
                "cbm": "1.1945",
                "unit_cost_usd": "46.00",   # from CI, may be None
            },
            ...
        ],
        "totals": {
            "cartons": 328,
            "pieces": 9535,
            "nw_kg": "6962.50",
            "gw_kg": "7971.60",
            "cbm": "41.8352",
        },
        "warnings": ["Could not find CI sheet — unit costs not imported"],
    }
"""

from __future__ import annotations

import re
from decimal import Decimal, InvalidOperation
from typing import Any


# ── helpers ──────────────────────────────────────────────────────────────────

def _safe_dec(val, places: int = 4) -> Decimal | None:
    if val is None or val == "":
        return None
    try:
        q = Decimal("0." + "0" * places)
        return Decimal(str(val)).quantize(q)
    except (InvalidOperation, ValueError):
        return None


def _dec2(val) -> Decimal | None:
    """Convenience wrapper — 2 decimal places for weight fields."""
    return _safe_dec(val, places=2)


def _safe_int(val) -> int | None:
    try:
        f = float(val)
        return int(round(f)) if f > 0 else None
    except (TypeError, ValueError):
        return None


def _norm(s: str) -> str:
    """Lowercase + collapse whitespace for fuzzy header matching."""
    return re.sub(r"\s+", " ", str(s)).strip().lower()


def _cell_str(sheet, row: int, col: int) -> str:
    try:
        return str(sheet.cell_value(row, col)).strip()
    except Exception:
        return ""


# ── sheet detectors ──────────────────────────────────────────────────────────

# Keywords we look for in header rows for each sheet type.
_PL_KEYWORDS = {"po", "p/n", "carton", "qty", "nw", "gw", "cbm"}
_CI_KEYWORDS  = {"unit cost", "unit price", "unit  cost", "单价"}


def _find_header_row(sheet, keywords: set[str], max_scan: int = 20) -> int | None:
    """Return the row index whose cells collectively contain all keywords."""
    for r in range(min(max_scan, sheet.nrows)):
        row_text = " ".join(_norm(_cell_str(sheet, r, c)) for c in range(sheet.ncols))
        if all(kw in row_text for kw in keywords):
            return r
    return None


def _col_idx(header_vals: list[str], *candidates: str) -> int | None:
    """Find the first column whose normalised header contains any candidate."""
    for candidate in candidates:
        for i, h in enumerate(header_vals):
            if candidate in _norm(h):
                return i
    return None


# ── packing list parser ───────────────────────────────────────────────────────

def _parse_pl_sheet(sheet) -> tuple[list[dict], dict, list[str]]:
    """
    Return (items, totals, warnings) from a packing-list-style sheet.

    Items contain everything except unit_cost_usd (filled later from CI).
    Totals are detected from the last summary row.
    """
    warnings: list[str] = []

    # Find the first header row that has PO + cartons + qty
    header_row = _find_header_row(sheet, {"po", "carton", "qty"})
    if header_row is None:
        warnings.append("Could not locate packing list header row.")
        return [], {}, warnings

    # There's often a second sub-header row (e.g. PCS/CTN / SUB / N.W/CTN / SUB …)
    # We want the main header, so skip subsequent sub-header rows.
    # Read the main header row columns.
    hdrs = [_cell_str(sheet, header_row, c) for c in range(sheet.ncols)]

    # Column mapping
    col_po    = _col_idx(hdrs, "po#", "po #", "po")
    col_pn    = _col_idx(hdrs, "p/n", "pn", "item no", "product no", "sku", "p/n")
    col_desc  = _col_idx(hdrs, "description", "product desc", "desc")
    col_ctn   = _col_idx(hdrs, "carton")
    col_qty   = _col_idx(hdrs, "qty.(pcs)", "qty(pcs)", "qty", "pcs", "quantity")
    col_cbm   = _col_idx(hdrs, "cbm")
    col_dims  = _col_idx(hdrs, "size", "measurement", "dimension", "meas")

    # NW and GW: look for sub-header row to find "SUB" columns
    # The main header may say "NW(kgs)" with merged cells above the sub-row.
    # Strategy: in the sub-header row (header_row+1), look for the second
    # "sub" column for GW total, and first "sub" for NW total.
    sub_row = header_row + 1
    nw_sub_col = None
    gw_sub_col = None
    if sub_row < sheet.nrows:
        sub_hdrs = [_norm(_cell_str(sheet, sub_row, c)) for c in range(sheet.ncols)]
        sub_indices = [i for i, h in enumerate(sub_hdrs) if "sub" in h]
        if len(sub_indices) >= 1:
            nw_sub_col = sub_indices[0]
        if len(sub_indices) >= 2:
            gw_sub_col = sub_indices[1]

        # CBM often only appears in the sub-header row — check there if not found above
        if col_cbm is None:
            col_cbm = _col_idx(sub_hdrs, "cbm")

    # Fallback: look for NW/GW columns in main header
    if nw_sub_col is None:
        nw_sub_col = _col_idx(hdrs, "nw(kgs)", "nw (kgs)", "nw(kg)", "n.w", "nw")
    if gw_sub_col is None:
        gw_sub_col = _col_idx(hdrs, "gw(kgs)", "gw (kgs)", "gw(kg)", "g.w", "gw")

    data_start = sub_row + 1 if sub_row < sheet.nrows and any(
        "sub" in _norm(_cell_str(sheet, sub_row, c)) for c in range(sheet.ncols)
    ) else header_row + 1

    items: list[dict] = []
    totals: dict = {}

    for r in range(data_start, sheet.nrows):
        # Skip blank rows
        row_vals = [_cell_str(sheet, r, c) for c in range(sheet.ncols)]
        if not any(v.strip() for v in row_vals):
            continue

        # Detect totals row: first col is blank and cartons col is a large-ish number
        po_val = row_vals[col_po] if col_po is not None else ""
        ctn_val = row_vals[col_ctn] if col_ctn is not None else ""

        # Totals row has no PO and a carton count in the right column
        if not po_val and ctn_val:
            try:
                tot_ctns = int(float(ctn_val))
                if tot_ctns > 1:
                    totals["cartons"] = tot_ctns
                    if col_qty is not None:
                        totals["pieces"] = _safe_int(row_vals[col_qty]) or _safe_int(
                            _cell_str(sheet, r, col_qty)
                        )
                    if nw_sub_col is not None:
                        totals["nw_kg"] = _dec2(row_vals[nw_sub_col])
                    if gw_sub_col is not None:
                        totals["gw_kg"] = _dec2(row_vals[gw_sub_col])
                    if col_cbm is not None:
                        totals["cbm"] = _safe_dec(row_vals[col_cbm])
            except (ValueError, TypeError):
                pass
            continue

        # Skip if no PO value and no P/N — likely a sub-total / note row
        pn_val = row_vals[col_pn] if col_pn is not None else ""
        if not po_val and not pn_val:
            continue

        item: dict[str, Any] = {
            "po_number":    po_val,
            "sku":          pn_val,
            "description":  (row_vals[col_desc].strip() if col_desc is not None else ""),
            "cartons":      _safe_int(ctn_val),
            "qty":          _safe_int(row_vals[col_qty]) if col_qty is not None else None,
            "nw_kg":        str(_dec2(row_vals[nw_sub_col])) if nw_sub_col is not None and _dec2(row_vals[nw_sub_col]) else None,
            "gw_kg":        str(_dec2(row_vals[gw_sub_col])) if gw_sub_col is not None and _dec2(row_vals[gw_sub_col]) else None,
            "dimensions_cm": (row_vals[col_dims].replace("*", "×") if col_dims is not None else ""),
            "cbm":          str(_safe_dec(row_vals[col_cbm])) if col_cbm is not None and _safe_dec(row_vals[col_cbm]) else None,
            "unit_cost_usd": None,
        }
        items.append(item)

    if not items:
        warnings.append("Packing list sheet found but no data rows extracted.")

    return items, totals, warnings


# ── CI parser ─────────────────────────────────────────────────────────────────

def _parse_ci_sheet(sheet) -> tuple[dict[tuple[str, str], Decimal], list[str]]:
    """
    Return ({(po_number, sku): unit_cost}, warnings) from a CI-style sheet.
    """
    warnings: list[str] = []

    header_row = _find_header_row(sheet, {"unit cost"})
    if header_row is None:
        # Try alternate keywords
        header_row = _find_header_row(sheet, {"unit price"})
    if header_row is None:
        warnings.append("Could not locate CI header row — unit costs not imported.")
        return {}, warnings

    hdrs = [_cell_str(sheet, header_row, c) for c in range(sheet.ncols)]
    col_po   = _col_idx(hdrs, "po#", "po #", "po")
    col_pn   = _col_idx(hdrs, "product no", "p/n", "item no", "sku")
    col_cost = _col_idx(hdrs, "unit cost", "unit price", "单价")

    if col_cost is None:
        warnings.append("Unit cost column not found in CI sheet.")
        return {}, warnings

    costs: dict[tuple[str, str], Decimal] = {}
    for r in range(header_row + 1, sheet.nrows):
        po  = _cell_str(sheet, r, col_po)  if col_po  is not None else ""
        pn  = _cell_str(sheet, r, col_pn)  if col_pn  is not None else ""
        raw = _cell_str(sheet, r, col_cost) if col_cost is not None else ""
        cost = _safe_dec(raw)
        if cost and (po or pn):
            costs[(po.strip(), pn.strip())] = cost

    return costs, warnings


# ── sheet classifier ──────────────────────────────────────────────────────────

def _classify_sheets(wb) -> tuple[Any | None, Any | None]:
    """Return (pl_sheet, ci_sheet) or None for each if not found."""
    pl_sheet = ci_sheet = None

    for name in wb.sheet_names():
        sh = wb.sheet_by_name(name)
        # Look at first 20 rows for keywords
        text = " ".join(
            _norm(_cell_str(sh, r, c))
            for r in range(min(20, sh.nrows))
            for c in range(sh.ncols)
        )
        has_pl = all(kw in text for kw in {"po", "carton", "qty", "cbm"})
        has_ci = any(kw in text for kw in {"unit cost", "unit price"}) and "po" in text

        if has_pl and pl_sheet is None:
            pl_sheet = sh
        if has_ci and ci_sheet is None:
            ci_sheet = sh

    return pl_sheet, ci_sheet


# ── public entry point ────────────────────────────────────────────────────────

def parse_shipment_doc(file_obj) -> dict:
    """
    Parse an uploaded file object (XLS or XLSX) and return structured shipment data.

    Accepts Django InMemoryUploadedFile / TemporaryUploadedFile.
    """
    import io

    warnings: list[str] = []

    # Read raw bytes
    file_obj.seek(0)
    raw = file_obj.read()
    filename = getattr(file_obj, "name", "").lower()

    # Try XLSX first, fall back to XLS
    wb = None
    if filename.endswith(".xlsx"):
        try:
            import openpyxl
            wb_xl = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)

            # Wrap openpyxl into an xlrd-like interface
            class _XlSheet:
                def __init__(self, ws):
                    self._ws = ws
                    self.nrows = ws.max_row or 0
                    self.ncols = ws.max_column or 0
                def cell_value(self, r, c):
                    v = self._ws.cell(row=r + 1, column=c + 1).value
                    return "" if v is None else v

            class _XlWb:
                def __init__(self, wb):
                    self._wb = wb
                def sheet_names(self):
                    return self._wb.sheetnames
                def sheet_by_name(self, name):
                    return _XlSheet(self._wb[name])

            wb = _XlWb(wb_xl)
        except ImportError:
            warnings.append("openpyxl not installed — trying xlrd for .xlsx.")

    if wb is None:
        try:
            import xlrd
            wb = xlrd.open_workbook(file_contents=raw)
        except Exception as e:
            return {"items": [], "totals": {}, "warnings": [f"Could not open file: {e}"]}

    pl_sheet, ci_sheet = _classify_sheets(wb)

    if pl_sheet is None:
        return {"items": [], "totals": {}, "warnings": ["No packing list sheet detected in this file."]}

    items, totals, pl_warnings = _parse_pl_sheet(pl_sheet)
    warnings.extend(pl_warnings)

    # Merge CI unit costs
    if ci_sheet is not None:
        costs, ci_warnings = _parse_ci_sheet(ci_sheet)
        warnings.extend(ci_warnings)
        for item in items:
            key = (item["po_number"], item["sku"])
            if key in costs:
                item["unit_cost_usd"] = str(costs[key])
            elif ("", item["sku"]) in costs:
                item["unit_cost_usd"] = str(costs[("", item["sku"])])
    else:
        warnings.append("No commercial invoice sheet detected — unit costs not imported.")

    # Convert Decimal totals to strings for JSON
    totals_out = {}
    for k, v in totals.items():
        totals_out[k] = str(v) if isinstance(v, Decimal) else v

    return {"items": items, "totals": totals_out, "warnings": warnings}
