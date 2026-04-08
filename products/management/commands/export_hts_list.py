"""
Management command: export_hts_list

Exports all HTS codes to an Excel spreadsheet for accounting.
Includes code, description, duty %, Section 301 %, extra tariff %,
total %, product count, and notes.

Usage:
    python manage.py export_hts_list
    python manage.py export_hts_list --output hts_list.xlsx
"""

import os

from django.core.management.base import BaseCommand
from django.db.models import Count

from products.models import HtsCode


class Command(BaseCommand):
    help = "Export all HTS codes to an Excel spreadsheet for accounting"

    def add_arguments(self, parser):
        parser.add_argument(
            "--output", default="hts_list.xlsx",
            help="Output filename (default: hts_list.xlsx)",
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        except ImportError:
            self.stderr.write(self.style.ERROR("openpyxl not installed. Run: pip install openpyxl"))
            return

        output_path = options["output"]

        codes = (
            HtsCode.objects
            .annotate(product_count=Count("products"))
            .order_by("code")
        )
        total = codes.count()
        self.stdout.write(f"Exporting {total} HTS codes...")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "HTS Codes"

        # ── Styles ────────────────────────────────────────────────────────
        header_fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
        header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        data_font   = Font(name="Arial", size=10)
        bold_font   = Font(name="Arial", size=10, bold=True)
        total_font  = Font(name="Arial", size=10, bold=True, color="1F4E79")
        center      = Alignment(horizontal="center", vertical="center")
        left        = Alignment(horizontal="left",   vertical="center")
        thin_bottom = Border(bottom=Side(style="thin", color="D9D9D9"))
        alt_fill    = PatternFill("solid", start_color="F2F7FB", end_color="F2F7FB")
        pct_num     = '0.00"%"'

        # ── Headers ───────────────────────────────────────────────────────
        headers = [
            "HTS Code", "Description",
            "Duty %", "Section 301 %", "Extra Tariff %", "Total %",
            "# Products", "Notes",
        ]
        col_widths = [16, 55, 10, 14, 14, 10, 12, 45]

        for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center
            ws.column_dimensions[cell.column_letter].width = w

        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

        # ── Data rows ─────────────────────────────────────────────────────
        for row_idx, hts in enumerate(codes, start=2):
            is_alt = (row_idx % 2 == 0)
            fill   = alt_fill if is_alt else None

            def cell(col, value, align=center, font=data_font, num_format=None):
                c = ws.cell(row=row_idx, column=col, value=value)
                c.font      = font
                c.alignment = align
                c.border    = thin_bottom
                if fill:
                    c.fill = fill
                if num_format:
                    c.number_format = num_format
                return c

            cell(1, hts.code,                          font=bold_font)
            cell(2, hts.description,                   align=left)
            cell(3, float(hts.duty_percent),           num_format=pct_num)
            cell(4, float(hts.section_301_percent),    num_format=pct_num)
            cell(5, float(hts.extra_tariff_percent),   num_format=pct_num)
            cell(6, float(hts.total_percent),          font=total_font, num_format=pct_num)
            cell(7, hts.product_count)
            cell(8, hts.other_tariff_notes or "",      align=left)

            ws.row_dimensions[row_idx].height = 16

        ws.auto_filter.ref = f"A1:H{total + 1}"

        wb.save(output_path)
        abs_path = os.path.abspath(output_path)
        self.stdout.write(self.style.SUCCESS(
            f"\n✓ Saved: {abs_path}\n"
            f"  {total} HTS codes exported"
        ))
