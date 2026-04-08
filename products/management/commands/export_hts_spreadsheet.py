"""
Management command: export_hts_spreadsheet

Exports all products to an Excel file for HTS code + category entry.

Sheet 1 — Products:
  SKU | Product Name | Current Category (scraped) | Correct Category | Sourcing | HTS Code
  Yellow = needs input.  Current Category shows what was scraped from the website
  (e.g. "Audio Tech") — enter the correct internal code in Correct Category.

Sheet 2 — Category Reference:
  Lists all Category codes and descriptions from the database.

Usage:
    python manage.py export_hts_spreadsheet
    python manage.py export_hts_spreadsheet --output hts_codes.xlsx
"""

import os

from django.core.management.base import BaseCommand

from products.models import Category, Product


class Command(BaseCommand):
    help = "Export all products to an Excel spreadsheet for HTS code + category entry"

    def add_arguments(self, parser):
        parser.add_argument(
            "--output",
            default="hts_codes.xlsx",
            help="Output filename (default: hts_codes.xlsx)",
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        except ImportError:
            self.stderr.write(self.style.ERROR(
                "openpyxl not installed. Run: pip install openpyxl"
            ))
            return

        output_path = options["output"]

        products = list(
            Product.objects
            .select_related("hts_code")
            .order_by("sku")
            .values("sku", "name", "category", "sourcing", "hts_code__code")
        )
        categories = list(Category.objects.order_by("code").values("code", "description"))

        total = len(products)
        self.stdout.write(f"Exporting {total} products, {len(categories)} categories...")

        wb = openpyxl.Workbook()

        # ── Shared styles ────────────────────────────────────────────────────
        header_fill  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
        header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        data_font    = Font(name="Arial", size=10)
        bold_font    = Font(name="Arial", size=10, bold=True)
        center       = Alignment(horizontal="center", vertical="center")
        left         = Alignment(horizontal="left",   vertical="center")
        thin_bottom  = Border(bottom=Side(style="thin", color="D9D9D9"))
        yellow_fill  = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
        alt_fill     = PatternFill("solid", start_color="F2F7FB", end_color="F2F7FB")
        note_fill    = PatternFill("solid", start_color="FCE4D6", end_color="FCE4D6")  # salmon — needs review

        SOURCING_LABELS = {"overseas": "Overseas", "domestic": "US Domestic", "retail": "Retail"}

        # ════════════════════════════════════════════════════════════════════
        # Sheet 1 — Products
        # ════════════════════════════════════════════════════════════════════
        ws = wb.active
        ws.title = "Products"

        headers    = ["SKU", "Product Name", "Current Category (scraped)", "Correct Category", "Sourcing", "HTS Code"]
        col_widths = [14,    52,              30,                           20,                  16,          22]

        for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center
            ws.column_dimensions[cell.column_letter].width = w

        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

        # Known website-scraped values that are NOT valid internal category codes
        SCRAPED_VALUES = {
            "Audio Tech", "Charging Tech", "Drinkware", "Lanyards",
            "Mobile Tech", "Office Tech", "Personal Tech", "USB Drives",
            "Other", "Accessories", "Apparel", "Bags", "Writing",
        }
        valid_codes = {c["code"] for c in categories}

        for row_idx, p in enumerate(products, start=2):
            is_alt = (row_idx % 2 == 0)

            current_cat = p["category"] or ""
            # Category is "wrong" if it's a scraped website value or not a valid code
            cat_is_scraped = current_cat in SCRAPED_VALUES or (
                current_cat != "" and current_cat not in valid_codes
            )

            def styled(col, value, align=left, font=data_font, fill=None):
                c = ws.cell(row=row_idx, column=col, value=value)
                c.font      = font
                c.alignment = align
                c.border    = thin_bottom
                if fill:
                    c.fill = fill
                elif is_alt:
                    c.fill = alt_fill

            # SKU
            styled(1, p["sku"], align=center, font=bold_font)

            # Name
            styled(2, p["name"])

            # Current Category (scraped) — salmon if it looks like a website nav value
            c = ws.cell(row=row_idx, column=3, value=current_cat)
            c.font      = data_font
            c.alignment = left
            c.border    = thin_bottom
            c.fill      = note_fill if cat_is_scraped else (alt_fill if is_alt else PatternFill())

            # Correct Category — yellow if needs fixing, pre-fill with existing if valid
            correct_cat = "" if cat_is_scraped else current_cat
            c = ws.cell(row=row_idx, column=4, value=correct_cat)
            c.font      = data_font
            c.alignment = center
            c.border    = thin_bottom
            c.fill      = yellow_fill if not correct_cat else (alt_fill if is_alt else PatternFill())

            # Sourcing
            styled(5, SOURCING_LABELS.get(p["sourcing"], p["sourcing"]), align=center)

            # HTS Code — yellow if blank
            existing_hts = p["hts_code__code"] or ""
            c = ws.cell(row=row_idx, column=6, value=existing_hts)
            c.font      = data_font
            c.alignment = center
            c.border    = thin_bottom
            c.fill      = yellow_fill if not existing_hts else (alt_fill if is_alt else PatternFill())

            ws.row_dimensions[row_idx].height = 16

        ws.auto_filter.ref = f"A1:F{total + 1}"

        # ════════════════════════════════════════════════════════════════════
        # Sheet 2 — Category Reference
        # ════════════════════════════════════════════════════════════════════
        ws2 = wb.create_sheet("Category Reference")

        ref_headers    = ["Code", "Description"]
        ref_col_widths = [14,     50]

        for col_idx, (h, w) in enumerate(zip(ref_headers, ref_col_widths), start=1):
            cell = ws2.cell(row=1, column=col_idx, value=h)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center
            ws2.column_dimensions[cell.column_letter].width = w

        ws2.row_dimensions[1].height = 22

        for row_idx, cat in enumerate(categories, start=2):
            c = ws2.cell(row=row_idx, column=1, value=cat["code"])
            c.font = bold_font; c.alignment = center; c.border = thin_bottom

            c = ws2.cell(row=row_idx, column=2, value=cat["description"])
            c.font = data_font; c.alignment = left; c.border = thin_bottom

            ws2.row_dimensions[row_idx].height = 16

        if not categories:
            ws2.cell(row=2, column=1, value="(No categories defined yet)").font = Font(
                name="Arial", size=10, italic=True, color="999999"
            )

        wb.save(output_path)
        abs_path = os.path.abspath(output_path)
        self.stdout.write(self.style.SUCCESS(
            f"\n✓ Saved: {abs_path}\n"
            f"  {total} products\n"
            f"  Salmon = scraped website category (needs correction)\n"
            f"  Yellow = blank field needing input\n"
            f"  See 'Category Reference' tab for valid category codes"
        ))
