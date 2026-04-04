"""
Management command: website_product_gap

Fetches all products from www.logoincluded.com/sitemap-3.xml,
compares with the ptool product database, and outputs an Excel
spreadsheet showing what's on the website vs. what's in ptool.

Usage:
    python manage.py website_product_gap
    python manage.py website_product_gap --output ~/Desktop/product_gap.xlsx
"""

import ssl
import urllib.request
import xml.etree.ElementTree as ET
from pathlib import Path

from django.core.management.base import BaseCommand

from products.models import Product


SITEMAP_URL = "https://www.logoincluded.com/sitemap-3.xml"


def fetch_website_products():
    """Parse logoincluded.com sitemap and return list of (sku, name) tuples."""
    req = urllib.request.Request(SITEMAP_URL, headers={"User-Agent": "Mozilla/5.0"})
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    with urllib.request.urlopen(req, timeout=15, context=ctx) as resp:
        xml_data = resp.read()

    root = ET.fromstring(xml_data)
    ns = {"sm": "http://www.sitemaps.org/schemas/sitemap/0.9"}
    locs = [el.text.strip() for el in root.findall(".//sm:loc", ns)]

    products = []
    for url in locs:
        if "/product/" in url:
            parts = url.split("/product/")[1]
            segments = parts.split("/")
            sku = segments[-1].upper()
            name = segments[0].replace("-", " ").title()
            products.append((sku, name))
    return products


class Command(BaseCommand):
    help = "Compare website products vs ptool database and export Excel gap report"

    def add_arguments(self, parser):
        parser.add_argument(
            "--output",
            default="product_gap_report.xlsx",
            help="Output Excel file path (default: product_gap_report.xlsx)",
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            self.stderr.write("openpyxl not installed. Run: pip install openpyxl")
            return

        output_path = Path(options["output"])
        self.stdout.write("Fetching website product list from sitemap...")

        try:
            website_products = fetch_website_products()
        except Exception as e:
            self.stderr.write(f"Failed to fetch sitemap: {e}")
            return

        self.stdout.write(f"Found {len(website_products)} products on website.")

        # Get all ptool SKUs
        ptool_products = list(Product.objects.values_list("sku", "name", "status"))
        ptool_skus = {row[0].upper(): (row[1], row[2]) for row in ptool_products}
        website_skus = {sku: name for sku, name in website_products}

        # Categorise
        website_only = [(sku, name) for sku, name in website_products if sku not in ptool_skus]
        in_both = [(sku, name) for sku, name in website_products if sku in ptool_skus]
        ptool_only = [(sku, ptool_skus[sku][0], ptool_skus[sku][1]) for sku in ptool_skus if sku not in website_skus]

        self.stdout.write(f"  On website only (not in ptool): {len(website_only)}")
        self.stdout.write(f"  In both website and ptool: {len(in_both)}")
        self.stdout.write(f"  In ptool only (not on website): {len(ptool_only)}")

        # --- Build Excel ---
        wb = openpyxl.Workbook()

        # Styles
        header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        header_fill_red = PatternFill("solid", fgColor="C0392B")
        header_fill_green = PatternFill("solid", fgColor="27AE60")
        header_fill_blue = PatternFill("solid", fgColor="2980B9")
        center = Alignment(horizontal="center", vertical="center")
        thin = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        alt_fill = PatternFill("solid", fgColor="F2F2F2")
        normal_font = Font(name="Arial", size=10)

        def write_sheet(ws, title, headers, rows, header_fill):
            ws.title = title
            # Header row
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
                cell.border = thin
            ws.row_dimensions[1].height = 20

            # Data rows
            for row_idx, row_data in enumerate(rows, 2):
                fill = alt_fill if row_idx % 2 == 0 else None
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = normal_font
                    cell.border = thin
                    cell.alignment = Alignment(vertical="center")
                    if fill:
                        cell.fill = fill

            # Auto-width
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

            # Freeze header
            ws.freeze_panes = "A2"

        # Sheet 1: Website only (not in ptool) — these are the legacy products
        ws1 = wb.active
        website_only_rows = [
            (sku, name, "", "", "", "")
            for sku, name in website_only
        ]
        write_sheet(
            ws1,
            "Not In Ptool",
            ["SKU", "Product Name (Website)", "Vendor", "Vendor SKU", "MOQ", "Notes"],
            website_only_rows,
            header_fill_red,
        )

        # Sheet 2: In both
        in_both_rows = [(sku, website_skus[sku], ptool_skus[sku][0], ptool_skus[sku][1]) for sku in [s for s, _ in in_both]]
        write_sheet(
            wb.create_sheet(),
            "In Both",
            ["SKU", "Website Name", "Ptool Name", "Ptool Status"],
            in_both_rows,
            header_fill_green,
        )

        # Sheet 3: Ptool only
        write_sheet(
            wb.create_sheet(),
            "Ptool Only",
            ["SKU", "Ptool Name", "Ptool Status"],
            ptool_only,
            header_fill_blue,
        )

        wb.save(output_path)
        self.stdout.write(self.style.SUCCESS(f"\nExcel report saved to: {output_path.resolve()}"))
        self.stdout.write(f"\nNext step: Fill in Vendor, Vendor SKU, and MOQ columns in the")
        self.stdout.write(f"'Not In Ptool' sheet and send back to Claude to import.")
